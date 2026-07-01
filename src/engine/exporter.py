import re
from pathlib import Path

from openpyxl import Workbook
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.cell.cell import Cell
from openpyxl.utils import get_column_letter
from src.models.testcase import Module
from src.core.config import ConfigManager

class ExcelExporter:
    def __init__(self, output_dir: str | Path):
        self.output_dir = Path(output_dir)
        self.output_dir.mkdir(parents=True, exist_ok=True)
        self.config = ConfigManager.get()

    def sanitize_sheet_name(self, name: str, existing_names: list[str]) -> str:
        # Max 28 chars, replace \ / * ? : [ ] with _
        clean_name = re.sub(r'[\\/\*\?\:\[\]]', '_', name)
        clean_name = clean_name[:28]
        
        final_name = clean_name
        counter = 1
        while final_name in existing_names:
            final_name = f"{clean_name[:26]}_{counter}"
            counter += 1
            
        existing_names.append(final_name)
        return final_name

    def parse_hyperlinks(self, text: str) -> tuple[object, str | None]:
        from openpyxl.cell.rich_text import CellRichText, TextBlock
        from openpyxl.cell.text import InlineFont

        matches = list(re.finditer(r'\[(.*?)\]\((.*?)\)', text))
        if not matches:
            return text, None
            
        font_link = InlineFont(u="single", color="0563C1")
        
        elements = []
        last_idx = 0
        first_url = None
        
        for match in matches:
            if first_url is None:
                first_url = match.group(2)
            
            if match.start() > last_idx:
                elements.append(text[last_idx:match.start()])
                
            elements.append(TextBlock(font_link, match.group(1)))
            last_idx = match.end()
            
        if last_idx < len(text):
            elements.append(text[last_idx:])
            
        return CellRichText(*elements), first_url

    def export(self, module: Module, output_filename: str):
        wb = Workbook()
        ws = wb.active
        assert isinstance(ws, Worksheet)
        
        sheet_name = self.sanitize_sheet_name(module.name, [])
        ws.title = sheet_name
        
        # Render Global Metadata
        row_idx = 1
        
        cell_module_label = ws.cell(row=row_idx, column=1, value="Module:")
        cell_module_label.font = self._bold_font()
        ws.cell(row=row_idx, column=2, value=module.name)
        row_idx += 1
        
        for key in self.config.global_metadata_keys:
            value = module.global_metadata.get(key, "")
            cell_key = ws.cell(row=row_idx, column=1, value=f"{key}:")
            cell_key.font = self._bold_font()
            ws.cell(row=row_idx, column=2, value=value)
            row_idx += 1
            
        row_idx += 1 # Empty row
        
        # Render Table Headers
        columns = sorted([c for c in self.config.columns if c.visible], key=lambda x: x.order)
        for col_idx, col in enumerate(columns, 1):
            cell_header = ws.cell(row=row_idx, column=col_idx, value=col.name)
            cell_header.font = self._bold_font()
            # Set column width
            ws.column_dimensions[get_column_letter(col_idx)].width = col.width
            
        row_idx += 1
        
        # Render Rows
        for scenario in module.scenarios:
            for col_idx, col in enumerate(columns, 1):
                cell = ws.cell(row=row_idx, column=col_idx)
                
                val = ""
                if col.id == "scenario":
                    val = scenario.name
                elif col.id == "precondition":
                    val = scenario.precondition
                elif col.id == "test_steps":
                    val = scenario.test_steps
                elif col.id == "expected_result":
                    val = scenario.expected_result
                elif col.name in scenario.local_metadata:
                    val = scenario.local_metadata[col.name]
                elif col.id in scenario.local_metadata:
                    val = scenario.local_metadata[col.id]
                    
                # pyrefly: ignore [unnecessary-type-conversion]
                display_text, url = self.parse_hyperlinks(str(val))
                # pyrefly: ignore [missing-attribute]
                cell.value = display_text
                if url:
                    cell.hyperlink = url  # type: ignore
                    
            row_idx += 1
            
        output_path = self.output_dir / output_filename
        wb.save(output_path)
        
    def _bold_font(self):
        from openpyxl.styles import Font
        return Font(bold=True)
